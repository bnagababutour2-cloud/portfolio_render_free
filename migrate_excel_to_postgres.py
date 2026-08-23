"""One-time migration from the existing Excel files to Postgres/Supabase.

Run locally on Windows, not on Render:
  set DATABASE_URL=...your Supabase connection string...
  python migrate_excel_to_postgres.py --portfolio E:\\PROJECT\\data.xlsm --clients E:\\PROJECT\\CLIENT_PHONENUMBER.xlsx

The script reads Excel only; it does not modify the source workbooks.
"""
import argparse
import hashlib
import hmac
import json
import os
import sqlite3
from pathlib import Path
from typing import Any

import psycopg
from openpyxl import load_workbook


def norm(v: Any) -> str:
    return "" if v is None else str(v).strip()


def clean_client_code(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    text = str(v).strip()
    try:
        f = float(text)
        if f.is_integer():
            return str(int(f))
    except Exception:
        pass
    return text


def number(v: Any):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except Exception:
        return v


def password_variants(value: Any) -> set[str]:
    result: set[str] = set()
    if value is None:
        return result
    if hasattr(value, "strftime"):
        result.update({
            value.strftime("%d-%m-%Y"), value.strftime("%d/%m/%Y"),
            value.strftime("%Y-%m-%d"), value.strftime("%d%m%Y"),
            value.strftime("%d.%m.%Y"),
        })
    else:
        text = str(value).strip()
        if text:
            result.add(text)
            result.add(text.upper())
            try:
                f = float(text)
                if f.is_integer():
                    result.add(str(int(f)))
            except Exception:
                pass
    return {x.strip().upper() for x in result if x and x.strip()}


def hash_password(value: str, secret: str) -> str:
    return hmac.new(secret.encode(), value.strip().upper().encode(), hashlib.sha256).hexdigest()


def first_existing(*paths: Path) -> Path:
    for p in paths:
        if p.is_file():
            return p
    return paths[0]


def normalize_headers(headers_raw):
    headers, seen = [], {}
    for i, h in enumerate(headers_raw):
        base = norm(h) or f"column_{i+1}"
        key = base.lower()
        seen[key] = seen.get(key, 0) + 1
        headers.append(base if seen[key] == 1 else f"{base}_{seen[key]}")
    return headers


def find_col(headers, *names):
    lookup = {h.strip().lower(): h for h in headers}
    for name in names:
        if name.lower() in lookup:
            return lookup[name.lower()]
    for h in headers:
        if any(name.lower() in h.lower() for name in names):
            return h
    return None


def prepare(record, headers):
    client_col = find_col(headers, "CLIENT CODE", "CLIENTCODE", "CLIENT")
    symbol_col = find_col(headers, "SYMBOL", "SCRIP", "STOCK")
    mtm_col = find_col(headers, "MTM")
    ltp_col = find_col(headers, "LTP")
    qty_col = find_col(headers, "BUY QUANTITY", "QUANTITY", "QTY")
    buy_col = find_col(headers, "BUY PRICE", "BUYPRICE")
    realised_col = find_col(headers, "REALISED PROFIT", "REALIZED PROFIT")
    change_col = find_col(headers, "CHANGE")
    return {
        "client_col": client_col, "symbol_col": symbol_col, "mtm_col": mtm_col,
        "ltp_col": ltp_col, "qty_col": qty_col, "buy_col": buy_col,
        "realised_col": realised_col, "change_col": change_col,
    }


def migrate(args):
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        raise SystemExit("DATABASE_URL is not set.")
    secret = os.environ.get("MIGRATION_PASSWORD_HASH_SECRET") or os.environ.get("PORTFOLIO_SESSION_SECRET")
    if not secret:
        raise SystemExit("Set MIGRATION_PASSWORD_HASH_SECRET or PORTFOLIO_SESSION_SECRET first.")

    portfolio_path = Path(args.portfolio)
    clients_path = Path(args.clients)
    if not portfolio_path.is_file():
        raise SystemExit(f"Portfolio workbook not found: {portfolio_path}")
    if not clients_path.is_file():
        raise SystemExit(f"Client workbook not found: {clients_path}")

    print("Reading portfolio workbook...")
    wb = load_workbook(portfolio_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers_raw = next(rows_iter, None)
    if not headers_raw:
        raise SystemExit("Portfolio workbook has no header row.")
    headers = normalize_headers(headers_raw)
    meta = prepare({}, headers)

    portfolio_rows = []
    for source_row, row in enumerate(rows_iter, start=2):
        padded = list(row) + [None] * max(0, len(headers) - len(row))
        data = {}
        for h, v in zip(headers, padded):
            data[h] = v.isoformat() if hasattr(v, "isoformat") else v
        client = clean_client_code(data.get(meta["client_col"])) if meta["client_col"] else ""
        symbol = norm(data.get(meta["symbol_col"])) if meta["symbol_col"] else ""
        if not client and not symbol:
            continue
        portfolio_rows.append((
            client, symbol,
            number(data.get(meta["mtm_col"])) if meta["mtm_col"] else None,
            number(data.get(meta["ltp_col"])) if meta["ltp_col"] else None,
            number(data.get(meta["qty_col"])) if meta["qty_col"] else None,
            number(data.get(meta["buy_col"])) if meta["buy_col"] else None,
            number(data.get(meta["realised_col"])) if meta["realised_col"] else None,
            number(data.get(meta["change_col"])) if meta["change_col"] else None,
            json.dumps(data, default=str),
            source_row,
        ))
    wb.close()

    print("Reading client credentials workbook...")
    cwb = load_workbook(clients_path, read_only=True, data_only=True)
    cws = cwb.active
    citer = cws.iter_rows(values_only=True)
    cheaders = next(citer, None)
    if not cheaders:
        raise SystemExit("Client workbook has no header row.")
    lookup = {str(h).strip().upper(): i for i, h in enumerate(cheaders) if h is not None}
    client_idx = lookup.get("CLIENT ID")
    mobile_idx = lookup.get("MOBILE")
    pan_idx = lookup.get("PAN NUMBR") or lookup.get("PAN NUMBER")
    dob_idx = lookup.get("DATE OF BIRTH")
    if client_idx is None:
        raise SystemExit("CLIENT ID column is missing from client workbook.")

    client_rows = {}
    for row in citer:
        if client_idx >= len(row):
            continue
        client = clean_client_code(row[client_idx])
        if not client:
            continue
        vals = []
        for idx in (mobile_idx, pan_idx, dob_idx):
            if idx is not None and idx < len(row):
                vals.extend(password_variants(row[idx]))
        hashes = sorted({hash_password(v, secret) for v in vals})
        client_rows[client] = hashes
    cwb.close()

    print(f"Connecting to database; importing {len(portfolio_rows)} portfolio rows and {len(client_rows)} clients...")
    with psycopg.connect(db_url) as con:
        with con.cursor() as cur:
            schema = Path(__file__).with_name("schema.sql").read_text(encoding="utf-8")
            cur.execute(schema)
            cur.execute("DELETE FROM portfolio")
            cur.execute("DELETE FROM clients")
            cur.execute("DELETE FROM app_meta WHERE key IN ('headers','meta')")
            cur.execute("INSERT INTO app_meta(key,value) VALUES (%s,%s),(%s,%s)",
                        ("headers", json.dumps(headers), "meta", json.dumps(meta)))

            for client, hashes in client_rows.items():
                cur.execute("INSERT INTO clients(client_id,password_hashes) VALUES (%s,%s)",
                            (client, json.dumps(hashes)))

            for row in portfolio_rows:
                client, symbol, mtm, ltp, qty, buy, realised, change_value, data_json, source_row = row
                cur.execute(
                    """INSERT INTO portfolio
                    (client_id,symbol,mtm,ltp,quantity,buy_price,realised_profit,change_value,data)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb)""",
                    (client, symbol, mtm, ltp, qty, buy, realised, change_value, data_json),
                )
        con.commit()

        if args.requests_db and Path(args.requests_db).is_file():
            print("Importing existing change request history...")
            with sqlite3.connect(args.requests_db) as old:
                old.row_factory = sqlite3.Row
                rows = old.execute("SELECT * FROM change_requests ORDER BY id").fetchall()
            with con.cursor() as cur:
                for r in rows:
                    keys = set(r.keys())
                    def g(k, default=None): return r[k] if k in keys else default
                    cur.execute(
                        """INSERT INTO change_requests
                        (client_id,symbol,excel_row,field,old_value,requested_value,requested_quantity,reason,status,created_at,reviewed_at,reviewed_by,admin_remark)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (g('client_id',''), g('symbol',''), g('excel_row'), g('field',''), g('old_value'),
                         g('requested_value',0), g('requested_quantity'), g('reason',''), g('status','PENDING'),
                         g('created_at',''), g('reviewed_at'), g('reviewed_by'), g('admin_remark'))
                    )
            con.commit()

    print("Migration complete.")
    print("IMPORTANT: verify a few client logins and portfolio rows before deleting your Excel backup.")


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--portfolio", required=True, help="Path to data.xlsm")
    p.add_argument("--clients", required=True, help="Path to CLIENT_PHONENUMBER.xlsx")
    p.add_argument("--requests-db", help="Optional path to portfolio_requests.db")
    migrate(p.parse_args())
